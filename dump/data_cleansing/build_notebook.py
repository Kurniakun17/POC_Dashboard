"""Generate the complete EDA.ipynb notebook for TEP Data Pipeline."""
import json

def md(source):
    return {"cell_type": "markdown", "metadata": {}, "source": source.split('\n'), "id": None}

def code(source):
    return {"cell_type": "code", "metadata": {}, "source": source.split('\n'), "outputs": [], "execution_count": None, "id": None}

cells = []

# ============================================================
# CELL 1: Title
# ============================================================
cells.append(md("""# BP Tangguh Expansion Project (TEP) - Data Pipeline & EDA
## EPCI Contract Analysis (2016-2024)

**Objective:** Extract, clean, normalize, and analyze TEP project data from 3 Excel sources into 13 relational database tables for MySQL import.

**Data Sources:**
1. `Contract Value Overview and Timeline_15-Jul-24.xlsx` - Monthly time-series (FGRS, LOGI, POB, Subcontractors)
2. `TEP Contract Evolution.xlsx` - Amendment-level cost breakdown (Original → AMD-5)
3. `Cost Breakdown Structure.xlsx` - PAMF claim hierarchy (3,967 claims / $1.77B)

**Target:** 6 Master Tables (`tb_m_*`) + 7 Transaction Tables (`tb_t_*`) → MySQL"""))

# ============================================================
# CELL 2: Imports
# ============================================================
cells.append(code("""import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns
from datetime import datetime
import os
import warnings
warnings.filterwarnings('ignore')

# Paths
RAW_DATA_DIR = r'D:\\BP\\raw_data'
OUTPUT_DIR = r'D:\\BP\\data_cleansing'

FILE_TIMELINE = os.path.join(RAW_DATA_DIR, 'Contract Value Overview and Timeline_15-Jul-24.xlsx')
FILE_EVOLUTION = os.path.join(RAW_DATA_DIR, 'TEP Contract Evolution.xlsx')
FILE_CBS = os.path.join(RAW_DATA_DIR, 'Cost Breakdown Structure.xlsx')

# Plot style
sns.set_theme(style='whitegrid', palette='deep')
plt.rcParams['figure.figsize'] = (14, 6)
plt.rcParams['figure.dpi'] = 100

PROJECT_ID = 1
print('Setup complete. Output directory:', OUTPUT_DIR)"""))

# ============================================================
# CELL 3: Section header
# ============================================================
cells.append(md("""---
## 1. RAW DATA EXTRACTION & STRUCTURE EXPLORATION
### 1.1 Load Excel Workbooks"""))

# ============================================================
# CELL 4: Load and inspect
# ============================================================
cells.append(code("""wb_timeline = openpyxl.load_workbook(FILE_TIMELINE, data_only=True)
wb_evolution = openpyxl.load_workbook(FILE_EVOLUTION, data_only=True)
wb_cbs = openpyxl.load_workbook(FILE_CBS, data_only=True)

print("=" * 70)
print("FILE 1: Contract Value Overview and Timeline")
print(f"  Sheets: {wb_timeline.sheetnames}")
for sn in wb_timeline.sheetnames:
    ws = wb_timeline[sn]
    print(f"    '{sn}': {ws.max_row} rows x {ws.max_column} cols")

print("\\nFILE 2: TEP Contract Evolution")
print(f"  Sheets: {wb_evolution.sheetnames}")
for sn in wb_evolution.sheetnames:
    ws = wb_evolution[sn]
    print(f"    '{sn}': {ws.max_row} rows x {ws.max_column} cols")

print("\\nFILE 3: Cost Breakdown Structure")
print(f"  Sheets: {wb_cbs.sheetnames}")
for sn in wb_cbs.sheetnames:
    ws = wb_cbs[sn]
    print(f"    '{sn}': {ws.max_row} rows x {ws.max_column} cols")"""))

# ============================================================
# CELL 5: Peek at raw data
# ============================================================
cells.append(code("""# Peek at first rows of each file
print("=== TEP Contract Evolution - First 10 rows ===")
ws_evo = wb_evolution[wb_evolution.sheetnames[0]]
for row in range(1, min(12, ws_evo.max_row + 1)):
    vals = [ws_evo.cell(row=row, column=c).value for c in range(1, min(10, ws_evo.max_column + 1))]
    print(f"  Row {row}: {vals}")

print("\\n=== Cost Breakdown Structure - First 10 rows ===")
ws_cbs_peek = wb_cbs[wb_cbs.sheetnames[0]]
for row in range(1, min(12, ws_cbs_peek.max_row + 1)):
    cell = ws_cbs_peek.cell(row=row, column=1)
    indent = cell.alignment.indent if cell.alignment and cell.alignment.indent else 0
    vals = [cell.value, ws_cbs_peek.cell(row=row, column=2).value, ws_cbs_peek.cell(row=row, column=3).value]
    print(f"  Row {row} (indent={indent}): {vals}")

print("\\n=== Timeline - Row 1 (years) ===")
ws_main = wb_timeline[wb_timeline.sheetnames[0]]
year_row = [ws_main.cell(row=1, column=c).value for c in range(1, min(90, ws_main.max_column + 1))]
print(f"  {[v for v in year_row if v is not None]}")
print("\\n=== Timeline - Row 3 (months) ===")
month_row = [ws_main.cell(row=3, column=c).value for c in range(1, min(90, ws_main.max_column + 1))]
print(f"  {[v for v in month_row if v is not None]}")
print("\\n=== Timeline - Column A+B labels (rows 1-60) ===")
for row in range(1, min(61, ws_main.max_row + 1)):
    a = ws_main.cell(row=row, column=1).value
    b = ws_main.cell(row=row, column=2).value
    c_val = ws_main.cell(row=row, column=3).value
    if a or b or c_val:
        print(f"  Row {row}: A={a}, B={b}, C={c_val}")"""))

# ============================================================
# CELL 6: Build column date map
# ============================================================
cells.append(md("""### 1.2 Build Year-Month Column Mapping for Timeline File
The timeline uses merged year headers (Row 1) and month numbers (Row 3). Years 2017/2018 are aggregated as "1-12", 2024 is quarterly."""))

cells.append(code("""ws_main = wb_timeline[wb_timeline.sheetnames[0]]

def build_column_date_map(ws):
    \"\"\"Map column indices to (year, month) tuples from merged header layout.\"\"\"
    # Forward-fill years from row 1
    year_map = {}
    current_year = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            try:
                current_year = int(v)
            except (ValueError, TypeError):
                pass
        if current_year:
            year_map[col] = current_year

    # Read months from row 3
    col_date_map = {}
    for col in range(1, ws.max_column + 1):
        month_val = ws.cell(row=3, column=col).value
        year = year_map.get(col)
        if month_val is None or year is None:
            continue
        if isinstance(month_val, (int, float)):
            col_date_map[col] = (year, int(month_val))
        elif isinstance(month_val, str):
            # Aggregated columns like '1 - 12', '1 - 3' etc
            col_date_map[col] = (year, month_val)
    return col_date_map

col_date_map = build_column_date_map(ws_main)

# Separate monthly vs aggregated
monthly_cols = {c: (y, m) for c, (y, m) in col_date_map.items() if isinstance(m, int)}
agg_cols = {c: (y, m) for c, (y, m) in col_date_map.items() if isinstance(m, str)}

print(f"Total column mappings: {len(col_date_map)}")
print(f"Monthly columns (usable): {len(monthly_cols)}")
print(f"Aggregated columns (skipped): {len(agg_cols)}")
if monthly_cols:
    print(f"Date range: {min(monthly_cols.values())} to {max(monthly_cols.values())}")
print(f"\\nAggregated columns: {agg_cols}")"""))

# ============================================================
# CELL 8: Time series helper
# ============================================================
cells.append(code("""def extract_time_series(ws, row_num, col_date_map, monthly_only=True):
    \"\"\"Extract numeric values from a row using column-date mapping.\"\"\"
    data = []
    for col, (year, month) in col_date_map.items():
        if monthly_only and not isinstance(month, int):
            continue
        val = ws.cell(row=row_num, column=col).value
        if val is not None and isinstance(val, (int, float)):
            data.append({'year': int(year), 'month': int(month) if isinstance(month, int) else None, 'value': float(val)})
    return data

print("Helper function extract_time_series() defined.")"""))

# ============================================================
# CELL 9: Extract Contract Evolution
# ============================================================
cells.append(md("""### 1.3 Extract TEP Contract Evolution Data"""))

cells.append(code("""ws_evo = wb_evolution[wb_evolution.sheetnames[0]]

# Detect column layout
print("Column headers:")
for col in range(1, ws_evo.max_column + 1):
    h1 = ws_evo.cell(row=1, column=col).value
    h2 = ws_evo.cell(row=2, column=col).value
    h3 = ws_evo.cell(row=3, column=col).value
    print(f"  Col {col} ({get_column_letter(col)}): row1={h1}, row2={h2}, row3={h3}")"""))

cells.append(code("""# Extract all data rows - adjust column mapping based on peek above
evolution_data = []
for row in range(1, ws_evo.max_row + 1):
    row_vals = {}
    for col in range(1, ws_evo.max_column + 1):
        row_vals[col] = ws_evo.cell(row=row, column=col).value
    evolution_data.append(row_vals)

df_evo_raw = pd.DataFrame(evolution_data)
print(f"Raw evolution data: {df_evo_raw.shape}")
df_evo_raw.head(50)"""))

cells.append(code("""# Structure the evolution data properly
# Identify the columns: typically B=row_no, C=description, D-I = Original through AMD-5
# Find the header row and data columns by checking content
header_row = None
for i, row_data in enumerate(evolution_data):
    vals = list(row_data.values())
    vals_str = [str(v).upper() if v else '' for v in vals]
    if any('ORIGINAL' in v or 'CONTRACT' in v for v in vals_str):
        print(f"Potential header at index {i}: {vals}")
        if header_row is None:
            header_row = i

# Build structured dataframe from evolution
# Try standard layout: col 2=number, col 3=description, cols 4-9=amendments
evo_structured = []
for row in range(1, ws_evo.max_row + 1):
    row_no = ws_evo.cell(row=row, column=2).value
    desc = ws_evo.cell(row=row, column=3).value
    if desc is None:
        continue
    vals = {}
    vals['row_no'] = row_no
    vals['description'] = str(desc).strip()
    vals['excel_row'] = row
    # Try columns 4-9 for amendment values
    for ci, amd_name in enumerate(['original_contract', 'amd_1', 'amd_2', 'amd_3', 'amd_4', 'amd_5']):
        v = ws_evo.cell(row=row, column=4 + ci).value
        vals[amd_name] = v
    evo_structured.append(vals)

df_evolution = pd.DataFrame(evo_structured)
print(f"Structured evolution: {df_evolution.shape}")
print("\\nDescriptions found:")
for _, r in df_evolution.iterrows():
    print(f"  [{r['row_no']}] {r['description']}: orig={r['original_contract']}")"""))

# ============================================================
# CELL: Extract PAMF
# ============================================================
cells.append(md("""### 1.4 Extract Cost Breakdown Structure (PAMF Claims)"""))

cells.append(code("""ws_cbs_data = wb_cbs[wb_cbs.sheetnames[0]]

pamf_data = []
current_l0 = None
current_l1 = None

for row in range(1, ws_cbs_data.max_row + 1):
    cell = ws_cbs_data.cell(row=row, column=1)
    label = cell.value
    indent = 0
    if cell.alignment and cell.alignment.indent:
        indent = int(cell.alignment.indent)
    count_val = ws_cbs_data.cell(row=row, column=2).value
    amount_val = ws_cbs_data.cell(row=row, column=3).value

    if label is None:
        continue
    label_str = str(label).strip()
    if label_str in ('', 'Grand Total', 'Row Labels', 'Count of PAMF', 'Sum of PAMF Claim Amount'):
        # Check if this is a header row
        if label_str == 'Grand Total' and count_val is not None:
            pamf_data.append({
                'discipline': 'TOTAL', 'category': None, 'subcategory': None,
                'label': label_str, 'level': -1,
                'claim_count': count_val, 'claim_amount_usd': amount_val,
                'excel_row': row
            })
        continue

    if indent == 0:
        current_l0 = label_str
        current_l1 = None
    elif indent == 1:
        current_l1 = label_str

    pamf_data.append({
        'discipline': current_l0 if indent >= 0 else None,
        'category': current_l1 if indent >= 1 else label_str if indent == 0 else None,
        'subcategory': label_str if indent == 2 else None,
        'label': label_str,
        'level': indent,
        'claim_count': count_val,
        'claim_amount_usd': amount_val,
        'excel_row': row
    })

df_pamf_raw = pd.DataFrame(pamf_data)
print(f"PAMF raw entries: {len(df_pamf_raw)}")
print(f"\\nBy level:")
print(df_pamf_raw.groupby('level')[['claim_count', 'claim_amount_usd']].sum())
print(f"\\nDisciplines (level 0):")
print(df_pamf_raw[df_pamf_raw['level'] == 0][['label', 'claim_count', 'claim_amount_usd']])"""))

# ============================================================
# CELL: Extract time series data
# ============================================================
cells.append(md("""### 1.5 Extract Time-Series Data from Timeline
Identify the correct rows for FGRS, LOGI, POB, and Subcontractor data."""))

cells.append(code("""# Scan row labels in column A, B, C to identify data rows
print("=== Timeline Row Labels (scanning cols A-C) ===")
row_labels = {}
for row in range(1, ws_main.max_row + 1):
    a = ws_main.cell(row=row, column=1).value
    b = ws_main.cell(row=row, column=2).value
    c = ws_main.cell(row=row, column=3).value
    label = a or b or c
    if label:
        label_str = str(label).strip()
        row_labels[row] = label_str
        # Check if this row has numeric data
        sample_vals = []
        for col in list(monthly_cols.keys())[:5]:
            v = ws_main.cell(row=row, column=col).value
            if v is not None and isinstance(v, (int, float)):
                sample_vals.append(v)
        has_data = "DATA" if sample_vals else "text"
        print(f"  Row {row:>3}: [{has_data:>4}] {label_str[:80]}  sample={sample_vals[:3]}")"""))

cells.append(code("""# Based on the scan above, identify key rows
# These may need adjustment based on actual file structure
# Let's search for specific keywords
fgrs_rows = [r for r, l in row_labels.items() if 'FGRS' in l.upper() or 'RCE' in l.upper()]
logi_rows = [r for r, l in row_labels.items() if 'LOGI' in l.upper()]
pob_rows = [r for r, l in row_labels.items() if 'POB' in l.upper()]
isolation_rows = [r for r, l in row_labels.items() if 'ISOLATION' in l.upper() or 'ISOL' in l.upper()]
meindo_rows = [r for r, l in row_labels.items() if 'MEINDO' in l.upper()]
penta_rows = [r for r, l in row_labels.items() if 'PENTA' in l.upper()]
daewoo_rows = [r for r, l in row_labels.items() if 'DAEWOO' in l.upper()]

print(f"FGRS-related rows: {fgrs_rows}")
print(f"LOGI-related rows: {logi_rows}")
print(f"POB-related rows: {pob_rows}")
print(f"Isolation rows: {isolation_rows}")
print(f"Meindo rows: {meindo_rows}")
print(f"Penta rows: {penta_rows}")
print(f"Daewoo rows: {daewoo_rows}")

# For each identified section, scan nearby rows for data
def find_data_rows(ws, start_row, num_rows=15):
    \"\"\"Scan rows near a label to find which have numeric data.\"\"\"
    results = []
    for r in range(max(1, start_row - 2), min(ws.max_row, start_row + num_rows)):
        label_a = ws.cell(row=r, column=1).value
        label_b = ws.cell(row=r, column=2).value
        label_c = ws.cell(row=r, column=3).value
        label = str(label_a or label_b or label_c or '').strip()
        data_count = 0
        for col in monthly_cols:
            v = ws.cell(row=r, column=col).value
            if v is not None and isinstance(v, (int, float)):
                data_count += 1
        if data_count > 0:
            results.append((r, label, data_count))
    return results

# Find data rows near each section
if meindo_rows:
    print("\\n=== Meindo section ===")
    for r, l, c in find_data_rows(ws_main, meindo_rows[0]):
        print(f"  Row {r}: '{l}' ({c} data points)")

if penta_rows:
    print("\\n=== Penta section ===")
    for r, l, c in find_data_rows(ws_main, penta_rows[0]):
        print(f"  Row {r}: '{l}' ({c} data points)")

if daewoo_rows:
    print("\\n=== Daewoo section ===")
    for r, l, c in find_data_rows(ws_main, daewoo_rows[0]):
        print(f"  Row {r}: '{l}' ({c} data points)")"""))

cells.append(code("""# Extract FGRS data - find the row with most monthly numeric data near FGRS label
# We'll use adaptive row detection based on the scan above
def find_best_data_row(ws, keyword_rows, col_date_map):
    \"\"\"Find the row with most numeric data near keyword rows.\"\"\"
    best_row = None
    best_count = 0
    for kr in keyword_rows:
        for r in range(max(1, kr - 2), min(ws.max_row, kr + 5)):
            count = 0
            for col in col_date_map:
                v = ws.cell(row=r, column=col).value
                if isinstance(v, (int, float)):
                    count += 1
            if count > best_count:
                best_count = count
                best_row = r
    return best_row, best_count

# FGRS Monthly Cost
fgrs_row, fgrs_count = find_best_data_row(ws_main, fgrs_rows if fgrs_rows else [43], col_date_map)
print(f"FGRS data row: {fgrs_row} ({fgrs_count} data points)")
fgrs_data = extract_time_series(ws_main, fgrs_row, col_date_map) if fgrs_row else []
df_fgrs = pd.DataFrame(fgrs_data)
if len(df_fgrs) > 0:
    df_fgrs['cost_type'] = 'FGRS_RCE'
    df_fgrs.rename(columns={'value': 'monthly_amount_musd'}, inplace=True)
    print(f"FGRS monthly data: {len(df_fgrs)} points")
    print(df_fgrs.head())
else:
    print("WARNING: No FGRS data found")"""))

cells.append(code("""# LOGI RCE Cumulative Cost
logi_row, logi_count = find_best_data_row(ws_main, logi_rows if logi_rows else [49], col_date_map)
print(f"LOGI data row: {logi_row} ({logi_count} data points)")
logi_data = extract_time_series(ws_main, logi_row, col_date_map) if logi_row else []
df_logi = pd.DataFrame(logi_data)
if len(df_logi) > 0:
    df_logi['cost_type'] = 'LOGI_RCE'
    df_logi.rename(columns={'value': 'monthly_amount_musd'}, inplace=True)
    # LOGI may be cumulative - check if monotonically increasing
    is_cumulative = df_logi['monthly_amount_musd'].is_monotonic_increasing
    print(f"LOGI data: {len(df_logi)} points (cumulative={is_cumulative})")
    if is_cumulative:
        df_logi['cumulative_amount_musd'] = df_logi['monthly_amount_musd']
        df_logi['monthly_amount_musd'] = df_logi['cumulative_amount_musd'].diff().fillna(df_logi['cumulative_amount_musd'].iloc[0])
        print("Converted cumulative to monthly by differencing")
    print(df_logi.head())
else:
    print("WARNING: No LOGI data found")"""))

cells.append(code("""# POB (Personnel on Board)
pob_row, pob_count = find_best_data_row(ws_main, pob_rows if pob_rows else [55], col_date_map)
print(f"POB data row: {pob_row} ({pob_count} data points)")
pob_data = extract_time_series(ws_main, pob_row, col_date_map) if pob_row else []
df_pob = pd.DataFrame(pob_data)
if len(df_pob) > 0:
    df_pob.rename(columns={'value': 'pob_count'}, inplace=True)
    print(f"POB data: {len(df_pob)} points, range: {df_pob['pob_count'].min():.0f} - {df_pob['pob_count'].max():.0f}")

# Isolation Facility
iso_row, iso_count = find_best_data_row(ws_main, isolation_rows if isolation_rows else [58], col_date_map)
print(f"\\nIsolation data row: {iso_row} ({iso_count} data points)")
iso_data = extract_time_series(ws_main, iso_row, col_date_map) if iso_row else []
df_isolation = pd.DataFrame(iso_data)
if len(df_isolation) > 0:
    df_isolation.rename(columns={'value': 'isolation_count'}, inplace=True)
    print(f"Isolation data: {len(df_isolation)} points")

# Merge POB and Isolation
if len(df_pob) > 0:
    if len(df_isolation) > 0:
        df_pob_combined = df_pob.merge(df_isolation, on=['year', 'month'], how='left')
        df_pob_combined['isolation_count'] = df_pob_combined['isolation_count'].fillna(0)
    else:
        df_pob_combined = df_pob.copy()
        df_pob_combined['isolation_count'] = 0
    print(f"\\nCombined POB data: {len(df_pob_combined)} rows")
    print(df_pob_combined.head())
else:
    df_pob_combined = pd.DataFrame(columns=['year', 'month', 'pob_count', 'isolation_count'])"""))

# ============================================================
# Subcontractor extraction
# ============================================================
cells.append(md("""### 1.6 Extract Subcontractor Monthly Data"""))

cells.append(code("""# Define subcontractor metric labels to search for
SUBCON_METRICS = [
    'achieved_manhours', 'invoice_value_musd', 'actual_vowd_musd',
    'pob_plan', 'pob_actual',
    'cumulative_progress_plan', 'cumulative_progress_actual',
    'monthly_progress_plan', 'monthly_progress_actual'
]

# Common metric keywords in Excel labels
METRIC_KEYWORDS = {
    'mhr': 'achieved_manhours', 'manhour': 'achieved_manhours', 'man hour': 'achieved_manhours',
    'invoice': 'invoice_value_musd', 'inv ': 'invoice_value_musd',
    'vowd': 'actual_vowd_musd', 'vow': 'actual_vowd_musd',
    'pob': 'pob_actual', 'manpower': 'pob_actual',
    'plan pob': 'pob_plan', 'plan manpower': 'pob_plan',
    'cum': 'cumulative_progress_actual', 'cumulative': 'cumulative_progress_actual',
    'plan cum': 'cumulative_progress_plan', 'plan cumul': 'cumulative_progress_plan',
    'monthly prog': 'monthly_progress_actual', 'month prog': 'monthly_progress_actual',
    'plan month': 'monthly_progress_plan',
}

def classify_metric(label):
    \"\"\"Classify a row label into a standard metric name.\"\"\"
    label_lower = label.lower()
    # Check most specific keywords first
    for kw in sorted(METRIC_KEYWORDS.keys(), key=len, reverse=True):
        if kw in label_lower:
            return METRIC_KEYWORDS[kw]
    return None

def extract_subcontractor_data(ws, header_row, num_scan_rows=15):
    \"\"\"Extract all metric rows for a subcontractor section.\"\"\"
    results = []
    for r in range(header_row, min(ws.max_row + 1, header_row + num_scan_rows)):
        label_a = ws.cell(row=r, column=1).value
        label_b = ws.cell(row=r, column=2).value
        label_c = ws.cell(row=r, column=3).value
        label = str(label_a or label_b or label_c or '').strip()
        if not label:
            continue

        # Count data points
        data = extract_time_series(ws, r, col_date_map)
        if len(data) == 0:
            continue

        metric = classify_metric(label)
        if metric is None:
            metric = f'unknown_{r}'

        for pt in data:
            pt['metric'] = metric
            pt['excel_row'] = r
            pt['raw_label'] = label
        results.extend(data)
    return results

# Extract each subcontractor
all_subcon_data = []
subcon_sections = {
    'Meindo': meindo_rows[0] if meindo_rows else None,
    'Penta': penta_rows[0] if penta_rows else None,
    'Daewoo': daewoo_rows[0] if daewoo_rows else None,
}

for name, start_row in subcon_sections.items():
    if start_row is None:
        print(f"WARNING: {name} section not found, skipping")
        continue
    data = extract_subcontractor_data(ws_main, start_row)
    for d in data:
        d['subcontractor'] = name
    all_subcon_data.extend(data)
    metrics_found = set(d['metric'] for d in data)
    print(f"{name} (start row {start_row}): {len(data)} data points, metrics: {metrics_found}")

df_subcon_raw = pd.DataFrame(all_subcon_data)
if len(df_subcon_raw) > 0:
    print(f"\\nTotal subcontractor records: {len(df_subcon_raw)}")
    print(df_subcon_raw.groupby(['subcontractor', 'metric']).size())
else:
    print("WARNING: No subcontractor data extracted")"""))

# ============================================================
# Save raw CSVs
# ============================================================
cells.append(md("""### 1.7 Save Raw Extracted Data"""))

cells.append(code("""raw_csvs = {
    'raw_contract_evolution.csv': df_evolution,
    'raw_pamf_claims.csv': df_pamf_raw,
}
if len(df_fgrs) > 0:
    raw_csvs['raw_fgrs_monthly.csv'] = df_fgrs
if len(df_logi) > 0:
    raw_csvs['raw_logi_monthly.csv'] = df_logi
if len(df_pob_combined) > 0:
    raw_csvs['raw_pob_monthly.csv'] = df_pob_combined
if len(df_subcon_raw) > 0:
    raw_csvs['raw_subcontractor_monthly.csv'] = df_subcon_raw

for filename, df in raw_csvs.items():
    path = os.path.join(OUTPUT_DIR, filename)
    df.to_csv(path, index=False)
    print(f"Saved: {filename} ({len(df)} rows)")

print(f"\\nRaw CSVs saved to: {OUTPUT_DIR}")"""))

# ============================================================
# SECTION 2: NORMALIZATION
# ============================================================
cells.append(md("""---
## 2. DATA NORMALIZATION & TABLE CREATION
### Build all 13 target tables from extracted data

### 2.1 Master Tables"""))

# tb_m_project
cells.append(code("""# === tb_m_project ===
df_tb_m_project = pd.DataFrame([{
    'project_id': 1,
    'project_code': 'TEP',
    'project_name': 'Tangguh Expansion Project',
    'client': 'BP Berau Ltd',
    'country': 'Indonesia',
    'contract_type': 'EPCI',
    'original_contract_value': 2432775726,
    'start_date': '2016-06-01',
    'planned_end_date': '2023-12-31',
    'actual_end_date': None,
    'status': 'In Progress'
}])
print("tb_m_project:")
print(df_tb_m_project.T)"""))

# tb_m_amendment
cells.append(code("""# === tb_m_amendment ===
# Extract revised contract price row from evolution
revised_row = df_evolution[df_evolution['description'].str.contains('Revised CONTRACT PRICE|REVISED CONTRACT|CONTRACT PRICE', case=False, na=False)]
if len(revised_row) == 0:
    # Try last row which typically has totals
    revised_row = df_evolution.iloc[[-1]]
    print(f"Using last row as total: {revised_row['description'].values}")

print(f"Contract price row: {revised_row[['description', 'original_contract', 'amd_1', 'amd_2', 'amd_3', 'amd_4', 'amd_5']].values}")

# Also extract category totals (A, B, C, D)
cat_a = df_evolution[df_evolution['description'].str.contains('EPC LUMP SUM|LUMP SUM', case=False, na=False)]
cat_b = df_evolution[df_evolution['description'].str.contains('REIMBURSABLE', case=False, na=False)]
cat_c = df_evolution[df_evolution['description'].str.contains('PROVISIONAL SUM', case=False, na=False)]
cat_d = df_evolution[df_evolution['description'].str.contains('BACKCHARGE', case=False, na=False)]

def safe_get(df, col, default=0):
    if len(df) > 0 and col in df.columns:
        v = df.iloc[0][col]
        return v if v is not None and not (isinstance(v, float) and np.isnan(v)) else default
    return default

amendments = []
amd_cols = ['original_contract', 'amd_1', 'amd_2', 'amd_3', 'amd_4', 'amd_5']
amd_codes = ['ORIGINAL', 'AMD-1', 'AMD-2', 'AMD-3', 'AMD-4', 'AMD-5']
amd_names = ['Original Contract', 'Amendment 1', 'Amendment 2', 'Amendment 3', 'Amendment 4', 'Amendment 5']
amd_dates = ['2016-06-01', '2017-01-01', '2019-05-01', '2020-12-01', '2022-06-01', '2024-01-01']
amd_remarks = [
    'Original EPCI Contract Award',
    'Novated contract adjustment',
    'Scope restructure: VOs 1-9, logistics RCE',
    'COVID-19 response, FGRS RCE',
    'COVID FM costs, expanded reimbursable',
    'Final: COVID Tier 4, PP35, commissioning'
]

for i, col in enumerate(amd_cols):
    amendments.append({
        'amendment_id': i + 1,
        'project_id': 1,
        'amendment_code': amd_codes[i],
        'amendment_name': amd_names[i],
        'effective_date': amd_dates[i],
        'total_contract_value': safe_get(revised_row, col),
        'lump_sum_value': safe_get(cat_a, col),
        'reimbursable_value': safe_get(cat_b, col),
        'provisional_sum_value': safe_get(cat_c, col),
        'backcharge_value': safe_get(cat_d, col),
        'remarks': amd_remarks[i]
    })

df_tb_m_amendment = pd.DataFrame(amendments)
print("\\ntb_m_amendment:")
print(df_tb_m_amendment[['amendment_code', 'total_contract_value', 'lump_sum_value', 'reimbursable_value', 'provisional_sum_value']].to_string())"""))

# tb_m_cost_category
cells.append(code("""# === tb_m_cost_category ===
# Build hierarchical cost categories from evolution data
cost_cats = []
cat_id = 0
current_parent = None
current_type = None

for _, row in df_evolution.iterrows():
    rn = str(row['row_no']) if row['row_no'] else ''
    desc = str(row['description']).strip()

    # Skip the total row
    if 'Revised CONTRACT PRICE' in desc or 'REVISED' in desc.upper():
        continue

    # Level 1: Main categories (A, B, C, D)
    if rn in ('A', 'B', 'C', 'D'):
        cat_id += 1
        current_parent = cat_id
        if rn == 'A':
            current_type = 'LUMP_SUM'
        elif rn == 'B':
            current_type = 'REIMBURSABLE'
        elif rn == 'C':
            current_type = 'PROVISIONAL'
        elif rn == 'D':
            current_type = 'BACKCHARGE'
        cost_cats.append({
            'category_id': cat_id,
            'parent_category_id': None,
            'category_code': rn,
            'category_name': desc,
            'category_type': current_type,
            'level': 1
        })
    elif rn and current_parent:
        # Sub-categories
        cat_id += 1
        cost_cats.append({
            'category_id': cat_id,
            'parent_category_id': current_parent,
            'category_code': str(rn),
            'category_name': desc,
            'category_type': current_type,
            'level': 2
        })

df_tb_m_cost_category = pd.DataFrame(cost_cats)
print(f"tb_m_cost_category: {len(df_tb_m_cost_category)} categories")
print(df_tb_m_cost_category[['category_id', 'parent_category_id', 'category_code', 'category_name', 'category_type', 'level']].to_string())"""))

# tb_m_cost_discipline
cells.append(code("""# === tb_m_cost_discipline ===
# Extract unique disciplines from PAMF data
pamf_disciplines = df_pamf_raw[df_pamf_raw['level'] == 0]['label'].unique().tolist()
pamf_categories = df_pamf_raw[df_pamf_raw['level'] == 1]['label'].unique().tolist()

# Map discipline groups
def get_discipline_group(name):
    name_upper = name.upper()
    if 'COVID' in name_upper:
        return 'COVID'
    elif 'LOGISTIC' in name_upper:
        return 'LOGISTIC'
    elif 'PMT' in name_upper:
        return 'PMT'
    elif 'SMT' in name_upper:
        return 'SMT'
    return 'OTHER'

disciplines = []
disc_id = 0
for d in pamf_disciplines:
    if d == 'TOTAL':
        continue
    disc_id += 1
    disciplines.append({
        'discipline_id': disc_id,
        'discipline_code': f'{disc_id:02d}',
        'discipline_name': d,
        'discipline_group': get_discipline_group(d)
    })

# Add sub-disciplines from PAMF level 1
for cat in pamf_categories:
    if cat and cat not in [d['discipline_name'] for d in disciplines]:
        disc_id += 1
        disciplines.append({
            'discipline_id': disc_id,
            'discipline_code': f'{disc_id:02d}',
            'discipline_name': cat,
            'discipline_group': 'SUB'
        })

df_tb_m_cost_discipline = pd.DataFrame(disciplines)
print(f"tb_m_cost_discipline: {len(df_tb_m_cost_discipline)} disciplines")
print(df_tb_m_cost_discipline.to_string())"""))

# tb_m_subcontractor
cells.append(code("""# === tb_m_subcontractor ===
df_tb_m_subcontractor = pd.DataFrame([
    {'subcontractor_id': 1, 'project_id': 1, 'subcontractor_name': 'Meindo',
     'contract_number': '', 'scope_of_work': 'Piping Erection',
     'contract_value': None, 'start_date': None, 'end_date': None},
    {'subcontractor_id': 2, 'project_id': 1, 'subcontractor_name': 'Penta',
     'contract_number': '', 'scope_of_work': 'Piping Erection',
     'contract_value': None, 'start_date': None, 'end_date': None},
    {'subcontractor_id': 3, 'project_id': 1, 'subcontractor_name': 'Daewoo',
     'contract_number': '1306754', 'scope_of_work': 'General Construction / Piping Erection',
     'contract_value': None, 'start_date': None, 'end_date': None},
])
print("tb_m_subcontractor:")
print(df_tb_m_subcontractor.to_string())"""))

# tb_m_event
cells.append(code("""# === tb_m_event ===
df_tb_m_event = pd.DataFrame([
    {'event_id': 1, 'project_id': 1, 'event_code': 'CONTRACT_AWARD',
     'event_name': 'Contract Award', 'event_type': 'MILESTONE',
     'start_date': '2016-06-01', 'end_date': None,
     'description': 'Original EPCI Contract signed'},
    {'event_id': 2, 'project_id': 1, 'event_code': 'AMD_1',
     'event_name': 'Amendment 1', 'event_type': 'AMENDMENT',
     'start_date': '2017-01-01', 'end_date': None,
     'description': 'Novated contract adjustment (GE contract)'},
    {'event_id': 3, 'project_id': 1, 'event_code': 'AMD_2',
     'event_name': 'Amendment 2', 'event_type': 'AMENDMENT',
     'start_date': '2019-05-01', 'end_date': None,
     'description': 'Major scope restructure: reimbursable, VOs, logistics'},
    {'event_id': 4, 'project_id': 1, 'event_code': 'COVID_PANDEMIC',
     'event_name': 'COVID-19 Pandemic Onset', 'event_type': 'PANDEMIC',
     'start_date': '2020-03-01', 'end_date': None,
     'description': 'COVID-19 begins affecting project, POB restricted 12000+ to ~6300'},
    {'event_id': 5, 'project_id': 1, 'event_code': 'FM',
     'event_name': 'Force Majeure Declaration', 'event_type': 'FORCE_MAJEURE',
     'start_date': '2020-12-01', 'end_date': None,
     'description': 'Force Majeure declared due to COVID-19 impact'},
    {'event_id': 6, 'project_id': 1, 'event_code': 'AMD_3',
     'event_name': 'Amendment 3', 'event_type': 'AMENDMENT',
     'start_date': '2020-12-01', 'end_date': None,
     'description': 'COVID response, FGRS RCE, additional provisional sums'},
    {'event_id': 7, 'project_id': 1, 'event_code': 'OUTBREAK_2',
     'event_name': '2nd COVID Outbreak (Delta)', 'event_type': 'OUTBREAK',
     'start_date': '2021-07-01', 'end_date': '2021-12-31',
     'description': 'Second COVID outbreak at Tangguh site (Delta variant)'},
    {'event_id': 8, 'project_id': 1, 'event_code': 'AMD_4',
     'event_name': 'Amendment 4', 'event_type': 'AMENDMENT',
     'start_date': '2022-06-01', 'end_date': None,
     'description': 'Extended COVID FM costs, expanded reimbursable'},
    {'event_id': 9, 'project_id': 1, 'event_code': 'OUTBREAK_3',
     'event_name': '3rd COVID Outbreak (Omicron)', 'event_type': 'OUTBREAK',
     'start_date': '2022-01-01', 'end_date': '2022-06-30',
     'description': 'Third COVID outbreak (Omicron variant)'},
    {'event_id': 10, 'project_id': 1, 'event_code': 'AMD_5',
     'event_name': 'Amendment 5', 'event_type': 'AMENDMENT',
     'start_date': '2024-01-01', 'end_date': None,
     'description': 'Final amendment: COVID Tier 4, labor law PP35, commissioning'},
])
print("tb_m_event:")
print(df_tb_m_event[['event_id', 'event_code', 'event_name', 'event_type', 'start_date']].to_string())"""))

# ============================================================
# SECTION 2.2: Transaction Tables
# ============================================================
cells.append(md("""### 2.2 Transaction Tables"""))

# tb_t_contract_value
cells.append(code("""# === tb_t_contract_value ===
# Melt evolution from wide to long format: one row per (amendment, line_item)
amd_col_map = {
    'original_contract': 1, 'amd_1': 2, 'amd_2': 3,
    'amd_3': 4, 'amd_4': 5, 'amd_5': 6
}

# Assign cost_category_id based on section hierarchy
df_evo_work = df_evolution.copy()
current_cat = None
categories = []
for _, row in df_evo_work.iterrows():
    rn = str(row['row_no']) if row['row_no'] else ''
    if rn == 'A':
        current_cat = 1
    elif rn == 'B':
        current_cat = 2
    elif rn == 'C':
        current_cat = 3
    elif rn == 'D':
        current_cat = 4
    categories.append(current_cat)
df_evo_work['cost_category_id'] = categories

# Skip the total row
df_evo_work = df_evo_work[~df_evo_work['description'].str.contains('Revised CONTRACT PRICE|REVISED CONTRACT', case=False, na=False)]

# Melt
melted = df_evo_work.melt(
    id_vars=['row_no', 'description', 'cost_category_id'],
    value_vars=list(amd_col_map.keys()),
    var_name='amendment_col', value_name='amount_usd'
)
melted['amendment_id'] = melted['amendment_col'].map(amd_col_map)
melted = melted.dropna(subset=['cost_category_id'])
# Keep rows even with null amounts (shows what was added in later amendments)

df_tb_t_contract_value = melted[['amendment_id', 'cost_category_id', 'description', 'amount_usd']].copy()
df_tb_t_contract_value.insert(0, 'id', range(1, len(df_tb_t_contract_value) + 1))
df_tb_t_contract_value['remarks'] = None

print(f"tb_t_contract_value: {len(df_tb_t_contract_value)} rows")
print(f"By amendment: {df_tb_t_contract_value.groupby('amendment_id').size().to_dict()}")
print(df_tb_t_contract_value.head(20).to_string())"""))

# tb_t_monthly_cost
cells.append(code("""# === tb_t_monthly_cost ===
cost_frames = []
if len(df_fgrs) > 0:
    cost_frames.append(df_fgrs[['year', 'month', 'cost_type', 'monthly_amount_musd']].copy())
if len(df_logi) > 0:
    logi_for_cost = df_logi[['year', 'month', 'cost_type', 'monthly_amount_musd']].copy()
    if 'cumulative_amount_musd' in df_logi.columns:
        logi_for_cost['cumulative_amount_musd'] = df_logi['cumulative_amount_musd'].values
    cost_frames.append(logi_for_cost)

if cost_frames:
    df_tb_t_monthly_cost = pd.concat(cost_frames, ignore_index=True)
    df_tb_t_monthly_cost.insert(0, 'id', range(1, len(df_tb_t_monthly_cost) + 1))
    df_tb_t_monthly_cost['project_id'] = 1
    # Add cumulative for FGRS
    fgrs_mask = df_tb_t_monthly_cost['cost_type'] == 'FGRS_RCE'
    if fgrs_mask.any():
        fgrs_sorted = df_tb_t_monthly_cost[fgrs_mask].sort_values(['year', 'month'])
        df_tb_t_monthly_cost.loc[fgrs_mask, 'cumulative_amount_musd'] = fgrs_sorted['monthly_amount_musd'].cumsum().values
    if 'cumulative_amount_musd' not in df_tb_t_monthly_cost.columns:
        df_tb_t_monthly_cost['cumulative_amount_musd'] = None
    print(f"tb_t_monthly_cost: {len(df_tb_t_monthly_cost)} rows")
    print(df_tb_t_monthly_cost.groupby('cost_type').size())
    print(df_tb_t_monthly_cost.head())
else:
    df_tb_t_monthly_cost = pd.DataFrame()
    print("WARNING: No monthly cost data")"""))

# tb_t_monthly_pob
cells.append(code("""# === tb_t_monthly_pob ===
if len(df_pob_combined) > 0:
    df_tb_t_monthly_pob = df_pob_combined.copy()
    df_tb_t_monthly_pob.insert(0, 'id', range(1, len(df_tb_t_monthly_pob) + 1))
    df_tb_t_monthly_pob['project_id'] = 1
    df_tb_t_monthly_pob['remarks'] = None
    print(f"tb_t_monthly_pob: {len(df_tb_t_monthly_pob)} rows")
    print(f"POB range: {df_tb_t_monthly_pob['pob_count'].min():.0f} - {df_tb_t_monthly_pob['pob_count'].max():.0f}")
    print(df_tb_t_monthly_pob.head())
else:
    df_tb_t_monthly_pob = pd.DataFrame()
    print("WARNING: No POB data")"""))

# tb_t_pamf_claim
cells.append(code("""# === tb_t_pamf_claim ===
# Use leaf-level (level 2) entries, plus level 0/1 summaries
pamf_for_table = df_pamf_raw[df_pamf_raw['level'].isin([0, 1, 2])].copy()
pamf_for_table = pamf_for_table[pamf_for_table['label'] != 'TOTAL']

# Map discipline to discipline_id
disc_map = dict(zip(df_tb_m_cost_discipline['discipline_name'], df_tb_m_cost_discipline['discipline_id']))

pamf_for_table['discipline_id'] = pamf_for_table['discipline'].map(disc_map)
pamf_for_table.insert(0, 'id', range(1, len(pamf_for_table) + 1))
pamf_for_table['project_id'] = 1

# Determine pamf_group
def get_pamf_group(row):
    d = str(row.get('discipline', '')).upper()
    if 'COVID' in d:
        return 'COVID'
    elif 'LOGISTIC' in d:
        return 'LOGISTIC'
    elif 'PMT' in d:
        return 'PMT'
    elif 'SMT' in d:
        return 'SMT'
    return 'OTHER'

pamf_for_table['pamf_group'] = pamf_for_table.apply(get_pamf_group, axis=1)

df_tb_t_pamf_claim = pamf_for_table[['id', 'project_id', 'discipline_id', 'discipline',
    'pamf_group', 'label', 'level', 'claim_count', 'claim_amount_usd']].copy()
df_tb_t_pamf_claim.rename(columns={'claim_count': 'pamf_count'}, inplace=True)

print(f"tb_t_pamf_claim: {len(df_tb_t_pamf_claim)} rows")
print(f"Total claim amount: ${df_tb_t_pamf_claim[df_tb_t_pamf_claim['level'] == 0]['claim_amount_usd'].sum():,.2f}")
print(df_tb_t_pamf_claim.head(10).to_string())"""))

# tb_t_variation_order
cells.append(code("""# === tb_t_variation_order ===
vo_rows = df_evolution[df_evolution['description'].str.contains('VO|Variation|variation', case=False, na=False)]
print(f"VO-related rows found: {len(vo_rows)}")
print(vo_rows[['row_no', 'description']].to_string())

vo_records = []
vo_id = 0
for _, row in vo_rows.iterrows():
    for amd_col, amd_id in amd_col_map.items():
        val = row[amd_col]
        if val is not None and isinstance(val, (int, float)) and not np.isnan(val):
            vo_id += 1
            vo_records.append({
                'vo_id': vo_id,
                'project_id': 1,
                'vo_number': str(row['row_no']) if row['row_no'] else '',
                'vo_name': row['description'],
                'amount_usd': val,
                'status': 'APPROVED',
                'approved_in_amendment': amd_id,
                'approved_date': None
            })

df_tb_t_variation_order = pd.DataFrame(vo_records)
print(f"\\ntb_t_variation_order: {len(df_tb_t_variation_order)} rows")
if len(df_tb_t_variation_order) > 0:
    print(df_tb_t_variation_order.to_string())"""))

# tb_t_subcontractor_monthly
cells.append(code("""# === tb_t_subcontractor_monthly ===
subcon_id_map = {'Meindo': 1, 'Penta': 2, 'Daewoo': 3}

if len(df_subcon_raw) > 0:
    df_tb_t_subcontractor_monthly = df_subcon_raw.copy()
    df_tb_t_subcontractor_monthly['subcontractor_id'] = df_tb_t_subcontractor_monthly['subcontractor'].map(subcon_id_map)
    df_tb_t_subcontractor_monthly.insert(0, 'id', range(1, len(df_tb_t_subcontractor_monthly) + 1))

    # Pivot to wide format per (subcontractor, year, month) with metric columns
    pivot_data = df_tb_t_subcontractor_monthly.pivot_table(
        index=['subcontractor_id', 'subcontractor', 'year', 'month'],
        columns='metric', values='value', aggfunc='first'
    ).reset_index()
    pivot_data.insert(0, 'id', range(1, len(pivot_data) + 1))

    print(f"tb_t_subcontractor_monthly (long): {len(df_tb_t_subcontractor_monthly)} rows")
    print(f"tb_t_subcontractor_monthly (wide/pivoted): {len(pivot_data)} rows")
    print(f"Columns: {list(pivot_data.columns)}")
    # Use long format for storage
else:
    df_tb_t_subcontractor_monthly = pd.DataFrame()
    pivot_data = pd.DataFrame()
    print("WARNING: No subcontractor monthly data")"""))

# tb_t_project_progress
cells.append(code("""# === tb_t_project_progress ===
if len(df_subcon_raw) > 0:
    # Extract progress metrics from subcontractor data
    progress_metrics = [m for m in df_subcon_raw['metric'].unique()
                       if 'progress' in m.lower() or 'cum' in m.lower()]
    print(f"Progress metrics found: {progress_metrics}")

    df_progress = df_subcon_raw[df_subcon_raw['metric'].isin(progress_metrics)].copy()

    if len(df_progress) > 0:
        # Pivot to get plan vs actual columns
        df_progress_pivot = df_progress.pivot_table(
            index=['subcontractor', 'year', 'month'],
            columns='metric', values='value', aggfunc='first'
        ).reset_index()

        df_progress_pivot['subcontractor_id'] = df_progress_pivot['subcontractor'].map(subcon_id_map)
        df_progress_pivot.insert(0, 'id', range(1, len(df_progress_pivot) + 1))
        df_progress_pivot['project_id'] = 1

        # Rename columns for clarity
        col_rename = {}
        for c in df_progress_pivot.columns:
            if 'plan' in str(c).lower() and 'cum' in str(c).lower():
                col_rename[c] = 'plan_progress_pct'
            elif 'actual' in str(c).lower() and 'cum' in str(c).lower():
                col_rename[c] = 'overall_progress_pct'
        df_progress_pivot.rename(columns=col_rename, inplace=True)

        df_tb_t_project_progress = df_progress_pivot
        print(f"tb_t_project_progress: {len(df_tb_t_project_progress)} rows")
        print(df_tb_t_project_progress.head())
    else:
        df_tb_t_project_progress = pd.DataFrame()
        print("No progress metrics found in subcontractor data")
else:
    df_tb_t_project_progress = pd.DataFrame()
    print("WARNING: No project progress data")"""))

# ============================================================
# SECTION 3: DATA CLEANSING
# ============================================================
cells.append(md("""---
## 3. DATA CLEANSING & VALIDATION"""))

cells.append(code("""# Collect all tables
all_tables = {
    'tb_m_project': df_tb_m_project,
    'tb_m_amendment': df_tb_m_amendment,
    'tb_m_cost_category': df_tb_m_cost_category,
    'tb_m_cost_discipline': df_tb_m_cost_discipline,
    'tb_m_subcontractor': df_tb_m_subcontractor,
    'tb_m_event': df_tb_m_event,
    'tb_t_contract_value': df_tb_t_contract_value,
    'tb_t_monthly_cost': df_tb_t_monthly_cost,
    'tb_t_monthly_pob': df_tb_t_monthly_pob,
    'tb_t_pamf_claim': df_tb_t_pamf_claim,
    'tb_t_variation_order': df_tb_t_variation_order,
    'tb_t_subcontractor_monthly': df_tb_t_subcontractor_monthly,
    'tb_t_project_progress': df_tb_t_project_progress,
}

# Remove empty tables
all_tables = {k: v for k, v in all_tables.items() if len(v) > 0}

print(f"{'Table':<35} {'Rows':>6} {'Cols':>5} {'Nulls':>8}")
print("=" * 60)
for name, df in all_tables.items():
    null_count = df.isnull().sum().sum()
    print(f"{name:<35} {len(df):>6} {len(df.columns):>5} {null_count:>8}")"""))

cells.append(code("""# === Data Cleansing Operations ===
print("=== Cleansing Operations ===\\n")

# 1. Strip whitespace from text columns
for tname, df in all_tables.items():
    for col in df.select_dtypes(include='object').columns:
        if hasattr(df[col], 'str'):
            all_tables[tname][col] = df[col].str.strip()
    print(f"[{tname}] Text columns stripped")

# 2. Standardize numeric columns
for tname in ['tb_t_monthly_cost']:
    if tname in all_tables:
        df = all_tables[tname]
        for col in df.select_dtypes(include='float64').columns:
            if 'amount' in col or 'musd' in col:
                df[col] = df[col].round(6)
        print(f"[{tname}] Financial columns rounded")

# 3. POB as integers
if 'tb_t_monthly_pob' in all_tables:
    df = all_tables['tb_t_monthly_pob']
    for col in ['pob_count', 'isolation_count']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    print("[tb_t_monthly_pob] POB counts cast to int")

# 4. Validate contract values are positive where expected
if 'tb_t_contract_value' in all_tables:
    df = all_tables['tb_t_contract_value']
    neg_vals = df[df['amount_usd'].apply(lambda x: isinstance(x, (int, float)) and x < 0)]
    if len(neg_vals) > 0:
        print(f"[tb_t_contract_value] Negative values found (expected for backcharges): {len(neg_vals)}")

# 5. Check PAMF claim amounts
if 'tb_t_pamf_claim' in all_tables:
    df = all_tables['tb_t_pamf_claim']
    total = df[df['level'] == 0]['claim_amount_usd'].sum()
    print(f"[tb_t_pamf_claim] Total claim (level 0): ${total:,.2f}")

print("\\nCleansing complete.")"""))

cells.append(code("""# === Referential Integrity Checks ===
print("=== Referential Integrity Checks ===\\n")

# tb_t_contract_value.amendment_id -> tb_m_amendment.amendment_id
if 'tb_t_contract_value' in all_tables and 'tb_m_amendment' in all_tables:
    cv_amds = set(all_tables['tb_t_contract_value']['amendment_id'].dropna().unique())
    m_amds = set(all_tables['tb_m_amendment']['amendment_id'].unique())
    orphaned = cv_amds - m_amds
    print(f"contract_value.amendment_id: {cv_amds}")
    print(f"  Valid: {cv_amds & m_amds}, Orphaned: {orphaned}")

# tb_t_subcontractor_monthly FK
if 'tb_t_subcontractor_monthly' in all_tables and 'tb_m_subcontractor' in all_tables:
    sc_ids = set(all_tables['tb_t_subcontractor_monthly']['subcontractor_id'].dropna().unique())
    m_sc = set(all_tables['tb_m_subcontractor']['subcontractor_id'].unique())
    print(f"\\nsubcontractor_monthly.subcontractor_id: {sc_ids}")
    print(f"  Valid: {sc_ids & m_sc}, Orphaned: {sc_ids - m_sc}")

# Cross-check amendment totals
if 'tb_m_amendment' in all_tables:
    print("\\nAmendment value summary:")
    for _, amd in all_tables['tb_m_amendment'].iterrows():
        print(f"  {amd['amendment_code']}: Total={amd['total_contract_value']:>15,.0f} | "
              f"LS={amd['lump_sum_value']:>15,.0f} | "
              f"Reimb={amd['reimbursable_value']:>15,.0f} | "
              f"Prov={amd['provisional_sum_value']:>15,.0f}")

print("\\nAll checks complete.")"""))

# ============================================================
# SECTION 4: EDA VISUALIZATIONS
# ============================================================
cells.append(md("""---
## 4. EXPLORATORY DATA ANALYSIS & VISUALIZATIONS
### 4.1 Contract Value Growth Analysis"""))

cells.append(code("""# Contract Value Evolution Bar Chart
if 'tb_m_amendment' in all_tables:
    fig, axes = plt.subplots(1, 2, figsize=(18, 7))

    df_amd = all_tables['tb_m_amendment']
    labels = df_amd['amendment_code'].values
    values = df_amd['total_contract_value'].values / 1e9

    colors = ['#2196F3', '#64B5F6', '#4CAF50', '#FF9800', '#F44336', '#9C27B0'][:len(labels)]
    bars = axes[0].bar(labels, values, color=colors, edgecolor='black', linewidth=0.5)
    axes[0].set_title('TEP Contract Value Evolution', fontsize=14, fontweight='bold')
    axes[0].set_ylabel('Contract Value (Billion USD)')
    axes[0].set_ylim(0, max(values) * 1.15)
    for bar, val in zip(bars, values):
        axes[0].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.05,
                     f'${val:.2f}B', ha='center', fontsize=9, fontweight='bold')

    # Stacked bar: cost category breakdown per amendment
    cat_data = {}
    for cat_name, cat_col in [('Lump Sum', 'lump_sum_value'), ('Reimbursable', 'reimbursable_value'),
                               ('Provisional', 'provisional_sum_value')]:
        cat_data[cat_name] = df_amd[cat_col].values / 1e9

    x = np.arange(len(labels))
    width = 0.6
    bottom = np.zeros(len(labels))
    cat_colors = ['#1565C0', '#43A047', '#EF6C00']

    for (cat_name, vals), color in zip(cat_data.items(), cat_colors):
        vals_safe = np.nan_to_num(vals, nan=0.0)
        axes[1].bar(x, vals_safe, width, bottom=bottom, label=cat_name, color=color, alpha=0.85)
        bottom += vals_safe

    axes[1].set_xticks(x)
    axes[1].set_xticklabels(labels)
    axes[1].set_title('Contract Composition by Cost Category', fontsize=14, fontweight='bold')
    axes[1].set_ylabel('Amount (Billion USD)')
    axes[1].legend(title='Category')

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, 'viz_01_contract_evolution.png'), dpi=150, bbox_inches='tight')
    plt.show()
    print("Saved: viz_01_contract_evolution.png")
else:
    print("Skipped: no amendment data")"""))

cells.append(md("""### 4.2 Cost Category Breakdown"""))

cells.append(code("""if 'tb_m_amendment' in all_tables and 'tb_t_pamf_claim' in all_tables:
    fig, axes = plt.subplots(1, 2, figsize=(16, 7))

    # AMD-5 (last amendment) breakdown donut
    df_amd = all_tables['tb_m_amendment']
    last_amd = df_amd.iloc[-1]
    cat_labels = ['Lump Sum', 'Reimbursable', 'Provisional']
    cat_vals = [abs(last_amd['lump_sum_value'] or 0),
                abs(last_amd['reimbursable_value'] or 0),
                abs(last_amd['provisional_sum_value'] or 0)]
    cat_vals_b = [v / 1e9 for v in cat_vals]

    if last_amd['backcharge_value'] and abs(last_amd['backcharge_value']) > 0:
        cat_labels.append('Backcharge')
        cat_vals_b.append(abs(last_amd['backcharge_value']) / 1e9)

    colors_pie = ['#1976D2', '#388E3C', '#F57C00', '#D32F2F'][:len(cat_labels)]
    wedges, texts, autotexts = axes[0].pie(cat_vals_b, labels=cat_labels, autopct='%1.1f%%',
        colors=colors_pie, startangle=90, pctdistance=0.85, wedgeprops=dict(width=0.4))
    total_b = sum(cat_vals_b)
    axes[0].set_title(f'{last_amd["amendment_code"]} Cost Split (${total_b:.2f}B)', fontsize=13, fontweight='bold')

    # PAMF claims by discipline
    df_pamf = all_tables['tb_t_pamf_claim']
    disc_summary = df_pamf[df_pamf['level'] == 0].groupby('pamf_group').agg(
        total=('claim_amount_usd', 'sum')
    ).sort_values('total', ascending=True)

    disc_colors = {'COVID': '#E53935', 'LOGISTIC': '#1E88E5', 'PMT': '#43A047', 'SMT': '#FB8C00', 'OTHER': '#757575'}
    bar_colors = [disc_colors.get(g, '#757575') for g in disc_summary.index]

    axes[1].barh(disc_summary.index, disc_summary['total'] / 1e6, color=bar_colors)
    axes[1].set_xlabel('Claim Amount (Million USD)')
    axes[1].set_title('PAMF Claims by Discipline Group', fontsize=13, fontweight='bold')
    for i, (idx, row) in enumerate(disc_summary.iterrows()):
        axes[1].text(row['total'] / 1e6 + 5, i, f'${row["total"]/1e6:.0f}M', va='center', fontsize=10)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, 'viz_02_cost_breakdown.png'), dpi=150, bbox_inches='tight')
    plt.show()
    print("Saved: viz_02_cost_breakdown.png")"""))

cells.append(md("""### 4.3 Monthly Cost Time-Series"""))

cells.append(code("""if 'tb_t_monthly_cost' in all_tables and len(all_tables['tb_t_monthly_cost']) > 0:
    fig, axes = plt.subplots(2, 1, figsize=(18, 10))

    df_mc = all_tables['tb_t_monthly_cost']

    # FGRS
    fgrs = df_mc[df_mc['cost_type'] == 'FGRS_RCE'].copy()
    if len(fgrs) > 0:
        fgrs['date'] = pd.to_datetime(fgrs.apply(lambda r: f"{int(r['year'])}-{int(r['month']):02d}-01", axis=1))
        fgrs = fgrs.sort_values('date')
        axes[0].fill_between(fgrs['date'], fgrs['monthly_amount_musd'], alpha=0.3, color='#1976D2')
        axes[0].plot(fgrs['date'], fgrs['monthly_amount_musd'], color='#1976D2', linewidth=2)
        # COVID shading
        axes[0].axvspan(pd.Timestamp('2020-03-01'), pd.Timestamp('2021-01-01'), alpha=0.08, color='red', label='COVID Wave 1')
        axes[0].axvspan(pd.Timestamp('2021-06-01'), pd.Timestamp('2022-01-01'), alpha=0.08, color='orange', label='Delta Outbreak')
        axes[0].set_title('FGRS RCE Monthly Cost (MUSD)', fontsize=13, fontweight='bold')
        axes[0].set_ylabel('Monthly Cost (MUSD)')
        axes[0].legend(loc='upper right')

    # LOGI
    logi = df_mc[df_mc['cost_type'] == 'LOGI_RCE'].copy()
    if len(logi) > 0:
        logi['date'] = pd.to_datetime(logi.apply(lambda r: f"{int(r['year'])}-{int(r['month']):02d}-01", axis=1))
        logi = logi.sort_values('date')
        if 'cumulative_amount_musd' in logi.columns:
            axes[1].fill_between(logi['date'], logi['cumulative_amount_musd'], alpha=0.3, color='#388E3C')
            axes[1].plot(logi['date'], logi['cumulative_amount_musd'], color='#388E3C', linewidth=2)
            axes[1].set_ylabel('Cumulative Cost (MUSD)')
        else:
            axes[1].plot(logi['date'], logi['monthly_amount_musd'], color='#388E3C', linewidth=2)
            axes[1].set_ylabel('Monthly Cost (MUSD)')
        axes[1].set_title('LOGI RCE Cost (MUSD)', fontsize=13, fontweight='bold')

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, 'viz_03_monthly_costs.png'), dpi=150, bbox_inches='tight')
    plt.show()
    print("Saved: viz_03_monthly_costs.png")
else:
    print("Skipped: no monthly cost data")"""))

cells.append(md("""### 4.4 Personnel on Board (POB) Analysis"""))

cells.append(code("""if 'tb_t_monthly_pob' in all_tables and len(all_tables['tb_t_monthly_pob']) > 0:
    fig, ax = plt.subplots(figsize=(18, 7))

    pob = all_tables['tb_t_monthly_pob'].copy()
    pob['date'] = pd.to_datetime(pob.apply(lambda r: f"{int(r['year'])}-{int(r['month']):02d}-01", axis=1))
    pob = pob.sort_values('date')

    ax.bar(pob['date'], pob['pob_count'], width=25, color='#1976D2', alpha=0.7, label='POB (CSTS+SubCon)')

    # Isolation overlay
    if 'isolation_count' in pob.columns:
        iso = pob[pob['isolation_count'] > 0]
        if len(iso) > 0:
            ax.bar(iso['date'], iso['isolation_count'], width=25, color='#E53935', alpha=0.7, label='Isolation Facility')

    ax.axhline(y=12000, color='gray', linestyle='--', alpha=0.5, label='Pre-COVID Peak (~12,000)')
    ax.axhline(y=6300, color='orange', linestyle='--', alpha=0.5, label='Restricted Level (~6,300)')

    # COVID shading
    ax.axvspan(pd.Timestamp('2020-03-01'), pd.Timestamp('2021-01-01'), alpha=0.06, color='red')
    ax.axvspan(pd.Timestamp('2021-06-01'), pd.Timestamp('2022-01-01'), alpha=0.06, color='orange')

    ax.set_title('Monthly Personnel on Board (POB) with COVID Impact', fontsize=14, fontweight='bold')
    ax.set_ylabel('Personnel Count')
    ax.legend(loc='upper right')

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, 'viz_04_pob_timeline.png'), dpi=150, bbox_inches='tight')
    plt.show()
    print("Saved: viz_04_pob_timeline.png")
else:
    print("Skipped: no POB data")"""))

cells.append(md("""### 4.5 Subcontractor Performance"""))

cells.append(code("""if 'tb_t_project_progress' in all_tables and len(all_tables['tb_t_project_progress']) > 0:
    subcontractors = all_tables['tb_t_project_progress']['subcontractor'].unique()
    n_sub = len(subcontractors)

    if n_sub > 0:
        fig, axes = plt.subplots(1, n_sub, figsize=(7 * n_sub, 7), sharey=True)
        if n_sub == 1:
            axes = [axes]

        for idx, (subcon, ax) in enumerate(zip(subcontractors, axes)):
            prog = all_tables['tb_t_project_progress'][
                all_tables['tb_t_project_progress']['subcontractor'] == subcon
            ].copy()
            prog['date'] = pd.to_datetime(prog.apply(lambda r: f"{int(r['year'])}-{int(r['month']):02d}-01", axis=1))
            prog = prog.sort_values('date')

            plan_col = 'plan_progress_pct' if 'plan_progress_pct' in prog.columns else None
            actual_col = 'overall_progress_pct' if 'overall_progress_pct' in prog.columns else None

            if plan_col and plan_col in prog.columns:
                plan_vals = prog[plan_col].dropna()
                if len(plan_vals) > 0:
                    ax.plot(prog.loc[plan_vals.index, 'date'], plan_vals * 100, 'b--', linewidth=2, label='Plan')

            if actual_col and actual_col in prog.columns:
                actual_vals = prog[actual_col].dropna()
                if len(actual_vals) > 0:
                    ax.plot(prog.loc[actual_vals.index, 'date'], actual_vals * 100, 'r-', linewidth=2, label='Actual')

            ax.set_title(f'{subcon} - Cumulative Progress', fontsize=12, fontweight='bold')
            if idx == 0:
                ax.set_ylabel('Progress (%)')
            ax.legend()
            ax.set_ylim(0, 105)
            ax.tick_params(axis='x', rotation=45)

        plt.suptitle('Subcontractor Progress S-Curves (Plan vs Actual)', fontsize=14, fontweight='bold', y=1.02)
        plt.tight_layout()
        plt.savefig(os.path.join(OUTPUT_DIR, 'viz_05_subcontractor_progress.png'), dpi=150, bbox_inches='tight')
        plt.show()
        print("Saved: viz_05_subcontractor_progress.png")
    else:
        print("No subcontractor data to plot")
else:
    print("Skipped: no project progress data")"""))

cells.append(md("""### 4.6 PAMF Claims Deep Dive"""))

cells.append(code("""if 'tb_t_pamf_claim' in all_tables and len(all_tables['tb_t_pamf_claim']) > 0:
    fig, axes = plt.subplots(1, 2, figsize=(18, 8))

    df_pamf = all_tables['tb_t_pamf_claim']

    # Top 15 sub-categories by claim amount (level 2 or lowest available)
    leaf_level = df_pamf['level'].max()
    top_claims = df_pamf[df_pamf['level'] == leaf_level].nlargest(15, 'claim_amount_usd')

    if len(top_claims) > 0:
        disc_colors = {'COVID': '#E53935', 'LOGISTIC': '#1E88E5', 'PMT': '#43A047', 'SMT': '#FB8C00', 'OTHER': '#757575'}
        bar_colors = [disc_colors.get(g, '#757575') for g in top_claims['pamf_group']]

        axes[0].barh(range(len(top_claims)), top_claims['claim_amount_usd'] / 1e6, color=bar_colors)
        axes[0].set_yticks(range(len(top_claims)))
        axes[0].set_yticklabels(top_claims['label'].str[:40], fontsize=8)
        axes[0].set_xlabel('Claim Amount (Million USD)')
        axes[0].set_title('Top 15 PAMF Claim Categories', fontsize=13, fontweight='bold')
        axes[0].invert_yaxis()

    # Count vs Amount scatter
    disc_data = df_pamf[df_pamf['level'] == 0].copy()
    if len(disc_data) > 0:
        scatter_colors = [disc_colors.get(g, '#757575') for g in disc_data['pamf_group']]
        axes[1].scatter(disc_data['pamf_count'], disc_data['claim_amount_usd'] / 1e6,
                       s=200, alpha=0.7, c=scatter_colors, edgecolors='black', linewidth=0.5)
        for _, row in disc_data.iterrows():
            axes[1].annotate(row['label'], (row['pamf_count'], row['claim_amount_usd'] / 1e6),
                           fontsize=10, fontweight='bold', ha='left',
                           xytext=(5, 5), textcoords='offset points')
        axes[1].set_xlabel('Number of Claims')
        axes[1].set_ylabel('Total Claim Amount (Million USD)')
        axes[1].set_title('PAMF: Claim Count vs Amount', fontsize=13, fontweight='bold')

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, 'viz_06_pamf_analysis.png'), dpi=150, bbox_inches='tight')
    plt.show()
    print("Saved: viz_06_pamf_analysis.png")
else:
    print("Skipped: no PAMF data")"""))

cells.append(md("""### 4.7 Statistical Summary"""))

cells.append(code("""print("=" * 70)
print("STATISTICAL SUMMARIES")
print("=" * 70)

if 'tb_t_monthly_cost' in all_tables and len(all_tables['tb_t_monthly_cost']) > 0:
    print("\\n=== Monthly Cost Statistics (MUSD) ===")
    print(all_tables['tb_t_monthly_cost'].groupby('cost_type')['monthly_amount_musd'].describe().round(3))

if 'tb_t_monthly_pob' in all_tables and len(all_tables['tb_t_monthly_pob']) > 0:
    print("\\n=== Monthly POB Statistics ===")
    print(all_tables['tb_t_monthly_pob'][['pob_count', 'isolation_count']].describe().round(0))

if 'tb_t_pamf_claim' in all_tables and len(all_tables['tb_t_pamf_claim']) > 0:
    print("\\n=== PAMF Claim Summary by Group ===")
    disc = all_tables['tb_t_pamf_claim'][all_tables['tb_t_pamf_claim']['level'] == 0].groupby('pamf_group').agg(
        total_claims=('pamf_count', 'sum'),
        total_amount_usd=('claim_amount_usd', 'sum'),
    ).round(2)
    disc['avg_per_claim'] = (disc['total_amount_usd'] / disc['total_claims']).round(2)
    print(disc)

if 'tb_m_amendment' in all_tables:
    print("\\n=== Amendment Value Growth ===")
    df_amd = all_tables['tb_m_amendment']
    orig = df_amd.iloc[0]['total_contract_value']
    for _, row in df_amd.iterrows():
        growth = ((row['total_contract_value'] - orig) / orig * 100) if orig else 0
        print(f"  {row['amendment_code']}: ${row['total_contract_value']:>15,.0f}  (+{growth:.1f}%)")"""))

# ============================================================
# SECTION 5: SAVE CLEANED CSVs
# ============================================================
cells.append(md("""---
## 5. SAVE CLEANED CSVs FOR ALL TABLES"""))

cells.append(code("""print("Saving cleaned CSVs...\\n")
for table_name, df in all_tables.items():
    filepath = os.path.join(OUTPUT_DIR, f'{table_name}.csv')
    df.to_csv(filepath, index=False)
    print(f"  {table_name}.csv  ({len(df)} rows, {len(df.columns)} cols)")

print(f"\\nAll {len(all_tables)} table CSVs saved to: {OUTPUT_DIR}")"""))

# ============================================================
# SECTION 6: MySQL SCRIPTS
# ============================================================
cells.append(md("""---
## 6. MySQL CREATE TABLE + INSERT SCRIPTS
### Generate DDL and DML for DBeaver import"""))

cells.append(code("""def get_mysql_type(col_name, dtype, df):
    \"\"\"Determine MySQL column type from pandas column.\"\"\"
    col_lower = col_name.lower()
    if col_lower.endswith('_id') and col_lower == df.columns[0].lower():
        return 'INT AUTO_INCREMENT'
    elif col_lower.endswith('_id'):
        return 'INT'
    elif 'date' in col_lower:
        return 'DATE'
    elif 'amount' in col_lower or 'value' in col_lower or 'usd' in col_lower:
        return 'DECIMAL(18,2)'
    elif 'progress' in col_lower or 'pct' in col_lower:
        return 'DECIMAL(10,6)'
    elif dtype in ('float64', 'float32'):
        return 'DECIMAL(18,6)'
    elif dtype in ('int64', 'int32'):
        return 'INT'
    elif col_lower in ('description', 'scope_of_work', 'remarks', 'scope'):
        return 'TEXT'
    else:
        if len(df) > 0:
            max_len = df[col_name].astype(str).str.len().max()
            return f'VARCHAR({max(int(max_len * 1.5), 50)})'
        return 'VARCHAR(255)'

def generate_create_table(table_name, df):
    \"\"\"Generate MySQL CREATE TABLE statement.\"\"\"
    lines = [f"DROP TABLE IF EXISTS `{table_name}`;"]
    lines.append(f"CREATE TABLE `{table_name}` (")
    pk_col = df.columns[0]
    col_defs = []
    for col in df.columns:
        mysql_type = get_mysql_type(col, str(df[col].dtype), df)
        nullable = '' if col == pk_col else ' NULL'
        col_defs.append(f"  `{col}` {mysql_type}{nullable}")
    col_defs.append(f"  PRIMARY KEY (`{pk_col}`)")
    lines.append(',\\n'.join(col_defs))
    lines.append(") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;")
    return '\\n'.join(lines)

def generate_inserts(table_name, df, batch_size=50):
    \"\"\"Generate MySQL INSERT statements with batching.\"\"\"
    if len(df) == 0:
        return f"-- No data for {table_name}"

    statements = []
    cols = ', '.join([f'`{c}`' for c in df.columns])

    for i in range(0, len(df), batch_size):
        batch = df.iloc[i:i+batch_size]
        values_list = []
        for _, row in batch.iterrows():
            vals = []
            for v in row.values:
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    vals.append('NULL')
                elif isinstance(v, str):
                    escaped = v.replace("\\\\", "\\\\\\\\").replace("'", "\\\\'")
                    vals.append(f"'{escaped}'")
                elif isinstance(v, (pd.Timestamp, datetime)):
                    vals.append(f"'{v.strftime('%Y-%m-%d')}'")
                else:
                    vals.append(str(v))
            values_list.append(f"  ({', '.join(vals)})")

        stmt = f"INSERT INTO `{table_name}` ({cols}) VALUES\\n" + ',\\n'.join(values_list) + ';'
        statements.append(stmt)

    return '\\n\\n'.join(statements)

print("SQL generator functions defined.")"""))

cells.append(code("""# Generate complete SQL script
all_sql = []
all_sql.append("-- ============================================================")
all_sql.append("-- TEP Data Pipeline - MySQL DDL + DML Script")
all_sql.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
all_sql.append(f"-- Tables: {len(all_tables)}")
all_sql.append("-- ============================================================")
all_sql.append("")
all_sql.append("SET FOREIGN_KEY_CHECKS = 0;")
all_sql.append("")

# DDL: CREATE TABLE statements
all_sql.append("-- ============================================================")
all_sql.append("-- DDL: CREATE TABLE STATEMENTS")
all_sql.append("-- ============================================================")
all_sql.append("")

# Order: master tables first, then transaction tables
table_order = [
    'tb_m_project', 'tb_m_amendment', 'tb_m_cost_category', 'tb_m_cost_discipline',
    'tb_m_subcontractor', 'tb_m_event',
    'tb_t_contract_value', 'tb_t_monthly_cost', 'tb_t_monthly_pob',
    'tb_t_pamf_claim', 'tb_t_variation_order', 'tb_t_subcontractor_monthly',
    'tb_t_project_progress'
]

for tname in table_order:
    if tname in all_tables:
        ddl = generate_create_table(tname, all_tables[tname])
        all_sql.append(ddl)
        all_sql.append("")

# DML: INSERT statements
all_sql.append("-- ============================================================")
all_sql.append("-- DML: INSERT STATEMENTS")
all_sql.append("-- ============================================================")
all_sql.append("")

for tname in table_order:
    if tname in all_tables:
        all_sql.append(f"-- {tname} ({len(all_tables[tname])} rows)")
        all_sql.append(generate_inserts(tname, all_tables[tname]))
        all_sql.append("")

all_sql.append("SET FOREIGN_KEY_CHECKS = 1;")
all_sql.append("")
all_sql.append("-- End of script")

# Save
sql_content = '\\n'.join(all_sql)
sql_path = os.path.join(OUTPUT_DIR, 'tep_mysql_insert.sql')
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write(sql_content)

print(f"MySQL script saved: {sql_path}")
print(f"Script size: {len(sql_content):,} characters")
print(f"Tables included: {sum(1 for t in table_order if t in all_tables)}")"""))

# ============================================================
# SECTION 7: DASHBOARD IDEAS
# ============================================================
cells.append(md("""---
## 7. DASHBOARD & VISUALIZATION RECOMMENDATIONS

### Dashboard 1: Executive Cost Overview
- **KPI Cards:** Original Contract ($2.43B), Current ($5.05B), Growth (+108%), Amendments (5)
- **Waterfall chart:** Incremental cost additions per amendment
- **Stacked area:** EPC Lump Sum vs Reimbursable vs Provisional over time
- **Timeline:** Contract milestones with key events annotated

### Dashboard 2: COVID-19 Impact Dashboard
- **Before/After POB:** 12,000+ vs 6,300 peak restriction
- **COVID provisional escalation:** $0 -> $236M -> $408M -> $420M across AMD-3/4/5
- **Isolation facility occupancy** timeline overlay
- **Schedule delay:** Original completion vs extended timeline
- **Cost by COVID type:** Direct, Other Related, Force Majeure, Tier 4

### Dashboard 3: Monthly Cost Burn Rate (FGRS + LOGI)
- **Dual-axis time series:** FGRS monthly + LOGI cumulative with budget lines
- **Cumulative S-curve** vs budget benchmarks per amendment period
- **Monthly burn rate heatmap** (months x years)
- **Budget utilization gauge:** FGRS $1,837M / LOGI $482M targets vs actuals

### Dashboard 4: Subcontractor Performance
- **Triple S-curve:** Meindo vs Penta vs Daewoo (plan vs actual)
- **POB efficiency:** Achieved manhours per person per month
- **Earned Value:** Invoice progression vs % complete
- **Schedule/Cost variance** monthly trend (SV, CV indicators)

### Dashboard 5: PAMF Claims Analysis
- **Treemap:** Claims by discipline > category > sub-category (area = amount)
- **Pareto chart:** Top 20 categories (80/20 rule visualization)
- **SMT dominance:** $1.46B = 82% of total claims breakdown
- **Scatter:** Claim count vs amount with outlier detection

### Dashboard 6: Reimbursable Cost Growth Tracker
- **Amendment-over-amendment growth:** $0 -> $438M -> $1.96B
- **Discipline breakdown:** Construction ($1.09B), Logistics ($447M), PM ($213M)
- **Post-FGRS tracking:** $188M -> $1.51B progression

**Recommended Tools:** Power BI, Tableau, or Plotly Dash for interactive filtering by amendment period, date range, subcontractor, and cost category."""))

# ============================================================
# SECTION 8: SUMMARY
# ============================================================
cells.append(md("""---
## 8. SUMMARY

### Output Files Generated

| File | Type | Description |
|------|------|-------------|
| `tb_m_project.csv` | Master | Project metadata |
| `tb_m_amendment.csv` | Master | 6 contract amendments |
| `tb_m_cost_category.csv` | Master | Hierarchical cost categories |
| `tb_m_cost_discipline.csv` | Master | PAMF discipline taxonomy |
| `tb_m_subcontractor.csv` | Master | 3 subcontractors |
| `tb_m_event.csv` | Master | Key project events |
| `tb_t_contract_value.csv` | Transaction | Contract values per amendment per line item |
| `tb_t_monthly_cost.csv` | Transaction | FGRS + LOGI monthly costs |
| `tb_t_monthly_pob.csv` | Transaction | Monthly POB + isolation counts |
| `tb_t_pamf_claim.csv` | Transaction | PAMF claim details |
| `tb_t_variation_order.csv` | Transaction | Variation orders |
| `tb_t_subcontractor_monthly.csv` | Transaction | Subcontractor monthly metrics |
| `tb_t_project_progress.csv` | Transaction | Plan vs actual progress |
| `tep_mysql_insert.sql` | SQL | Complete CREATE TABLE + INSERT statements for DBeaver |
| `raw_*.csv` | Raw | Raw extraction CSVs |
| `viz_*.png` | Charts | EDA visualization images |"""))

cells.append(code("""# Final summary
print("=" * 70)
print("TEP DATA PIPELINE - EXECUTION COMPLETE")
print("=" * 70)
print(f"\\nOutput directory: {OUTPUT_DIR}")
print(f"\\nTables generated:")
for name, df in all_tables.items():
    print(f"  {name}: {len(df)} rows x {len(df.columns)} cols")
print(f"\\nSQL script: tep_mysql_insert.sql")
print(f"\\nVisualization files: viz_01 through viz_06")

# List all output files
import glob
output_files = glob.glob(os.path.join(OUTPUT_DIR, '*'))
print(f"\\nAll output files ({len(output_files)}):")
for f in sorted(output_files):
    size = os.path.getsize(f)
    print(f"  {os.path.basename(f)}: {size:,} bytes")"""))

# ============================================================
# Build the notebook JSON
# ============================================================
# Fix cell sources to be proper list of lines
for i, cell in enumerate(cells):
    if isinstance(cell['source'], list):
        # Already a list, join and re-split properly
        text = '\n'.join(cell['source'])
    else:
        text = cell['source']
    # Split into lines, each ending with \n except the last
    lines = text.split('\n')
    cell['source'] = [line + '\n' for line in lines[:-1]] + [lines[-1]]
    cell['id'] = f'cell_{i:03d}'

notebook = {
    "nbformat": 4,
    "nbformat_minor": 5,
    "metadata": {
        "kernelspec": {
            "display_name": "Python 3",
            "language": "python",
            "name": "python3"
        },
        "language_info": {
            "name": "python",
            "version": "3.11.0"
        }
    },
    "cells": cells
}

output_path = r'D:\BP\data_cleansing\EDA.ipynb'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(notebook, f, indent=1, ensure_ascii=False)

print(f"Notebook written: {output_path}")
print(f"Total cells: {len(cells)}")
print(f"  Markdown: {sum(1 for c in cells if c['cell_type'] == 'markdown')}")
print(f"  Code: {sum(1 for c in cells if c['cell_type'] == 'code')}")
